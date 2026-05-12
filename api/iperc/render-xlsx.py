from http.server import BaseHTTPRequestHandler
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from io import BytesIO
from email.parser import BytesParser
from email.policy import default
import json


DATA_START_ROW = 21


def reemplazar_placeholders(ws, cabecera: dict):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                nueva = cell.value
                for k, v in cabecera.items():
                    placeholder = '{{' + k + '}}'
                    if placeholder in nueva:
                        nueva = nueva.replace(placeholder, str(v) if v is not None else '')
                cell.value = nueva


def aplicar_borde(cell):
    thin = Side(border_style='thin', color='000000')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def insertar_filas_iperc(ws, filas: list):
    for i, fila in enumerate(filas):
        r = DATA_START_ROW + i

        ws.cell(r, 2).value = fila.get('numero', i + 1)
        ws.cell(r, 3).value = fila.get('proceso', '')
        ws.cell(r, 4).value = fila.get('actividad', '')
        ws.cell(r, 5).value = fila.get('tarea', '')
        ws.cell(r, 6).value = fila.get('puestoTrabajo', '')
        ws.cell(r, 7).value = fila.get('factorRiesgo', '')
        ws.cell(r, 8).value = fila.get('condicionActividad', '')
        ws.cell(r, 9).value = fila.get('peligro', '')
        ws.cell(r, 10).value = fila.get('riesgo', '')
        ws.cell(r, 11).value = fila.get('consecuencia', '')
        ws.cell(r, 12).value = fila.get('severidad', '')
        ws.cell(r, 13).value = fila.get('probabilidad', '')
        ws.cell(r, 14).value = f'=L{r}&M{r}'
        ws.cell(r, 15).value = f'=IFERROR(VLOOKUP(N{r},Matriz!$A:$D,4,FALSE),"")'
        ws.cell(r, 16).value = f'=IFERROR(VLOOKUP(N{r},Matriz!$A:$E,5,FALSE),"")'

        ws.cell(r, 17).value = fila.get('eliminar', 'NA')
        ws.cell(r, 18).value = fila.get('sustituir', 'NA')
        ws.cell(r, 19).value = fila.get('controlIngenieria', '')
        ws.cell(r, 20).value = fila.get('controlAdministrativo', '')
        ws.cell(r, 21).value = fila.get('controlReceptor', '')

        ws.cell(r, 22).value = fila.get('severidadResidual', '')
        ws.cell(r, 23).value = fila.get('probabilidadResidual', '')
        ws.cell(r, 24).value = f'=V{r}&W{r}'
        ws.cell(r, 25).value = f'=IFERROR(VLOOKUP(X{r},Matriz!$A:$D,4,FALSE),"")'
        ws.cell(r, 26).value = f'=IFERROR(VLOOKUP(X{r},Matriz!$A:$E,5,FALSE),"")'

        ws.cell(r, 27).value = fila.get('accionesMejora', '')
        ws.cell(r, 28).value = fila.get('responsables', '')

        for col in range(2, 29):
            c = ws.cell(r, col)
            aplicar_borde(c)
            c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')


def parsear_multipart(body: bytes, content_type: str) -> dict:
    parser_input = b'Content-Type: ' + content_type.encode() + b'\r\n\r\n' + body
    msg = BytesParser(policy=default).parsebytes(parser_input)
    parts = {}
    for part in msg.iter_parts():
        nombre = part.get_param('name', header='content-disposition')
        if nombre:
            parts[nombre] = part.get_payload(decode=True)
    return parts


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            content_type = self.headers.get('Content-Type', '')
            body = self.rfile.read(content_length)

            parts = parsear_multipart(body, content_type)
            plantilla_bytes = parts.get('plantilla')
            data_bag_raw = parts.get('dataBag')

            if not plantilla_bytes or not data_bag_raw:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({'error': 'Faltan plantilla o dataBag'}).encode())
                return

            data_bag = json.loads(data_bag_raw.decode('utf-8'))
            cabecera = data_bag.get('cabecera', {})
            filas = data_bag.get('filas', [])

            wb = load_workbook(BytesIO(plantilla_bytes))

            ws_name = 'IPERC'
            if ws_name not in wb.sheetnames:
                raise ValueError(f'Hoja "{ws_name}" no encontrada. Hojas disponibles: {wb.sheetnames}')
            ws = wb[ws_name]

            reemplazar_placeholders(ws, cabecera)
            insertar_filas_iperc(ws, filas)

            output = BytesIO()
            wb.save(output)
            output_bytes = output.getvalue()

            self.send_response(200)
            self.send_header(
                'Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            self.send_header('Content-Length', str(len(output_bytes)))
            self.end_headers()
            self.wfile.write(output_bytes)

        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({
                'error': str(e),
                'type': type(e).__name__
            }).encode('utf-8'))
