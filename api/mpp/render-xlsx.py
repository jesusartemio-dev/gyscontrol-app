from http.server import BaseHTTPRequestHandler
from openpyxl import load_workbook
from io import BytesIO
from email.parser import BytesParser
from email.policy import default
import json

SHEET_NAME = 'MATRIZ EPPs'
DATA_START_ROW = 9

# Encabezados de puesto: fila 8, columnas H-Q (8-17), individuales (no
# combinadas). La plantilla trae hasta 10 slots — proyectos con más puestos
# se capan a 10 (ver do_POST). Las columnas sobrantes se blanquean.
FILA_HEADER_PUESTOS = 8
COL_PUESTOS_INICIO = 8
COL_PUESTOS_FIN = 17
MAX_PUESTOS = COL_PUESTOS_FIN - COL_PUESTOS_INICIO + 1

COL_NUMERO = 2
COL_NOMBRE = 3
# COL 4 = EVIDENCIA — NO tocar, contiene imágenes EPP
COL_RIESGO = 5
COL_PARTE = 6
COL_DURABILIDAD = 7


def reemplazar_placeholders(ws, cabecera: dict):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '{{' in cell.value:
                nueva = cell.value
                for k, v in cabecera.items():
                    placeholder = '{{' + k + '}}'
                    if placeholder in nueva:
                        nueva = nueva.replace(placeholder, str(v) if v else '')
                cell.value = nueva


def escribir_encabezados_puestos(ws, puestos: list) -> list:
    """Escribe los cargos del proyecto en la fila 8 (H8:Q8), capado a
    MAX_PUESTOS. openpyxl solo cambia el .value de la celda — preserva el
    estilo/rotación ya dibujados en la plantilla. Blanquea las columnas
    sobrantes (proyectos con menos de 10 puestos). Devuelve la lista
    [(puesto, columna), ...] efectivamente escrita."""
    puestos_capados = puestos[:MAX_PUESTOS]
    columnas = []
    for i in range(MAX_PUESTOS):
        col = COL_PUESTOS_INICIO + i
        if i < len(puestos_capados):
            ws.cell(FILA_HEADER_PUESTOS, col).value = puestos_capados[i]
            columnas.append((puestos_capados[i], col))
        else:
            ws.cell(FILA_HEADER_PUESTOS, col).value = None
    return columnas


def escribir_filas_mpp(ws, items: list, columnas_puestos: list):
    columnas_usadas = {col for _, col in columnas_puestos}
    for i, item in enumerate(items):
        r = DATA_START_ROW + i

        ws.cell(r, COL_NUMERO).value = i + 1
        ws.cell(r, COL_NOMBRE).value = item.get('nombre', '')
        ws.cell(r, COL_RIESGO).value = item.get('riesgo', '')
        ws.cell(r, COL_PARTE).value = item.get('parteCuerpo', '')
        ws.cell(r, COL_DURABILIDAD).value = item.get('durabilidad', '')
        # Columna 4 (EVIDENCIA) NO se toca — imágenes EPP están ahí

        asignaciones = item.get('asignaciones', {})
        for puesto, col in columnas_puestos:
            ws.cell(r, col).value = 'X' if asignaciones.get(puesto) is True else None
        # Blanquear columnas de puestos sobrantes (proyectos con < 10 puestos)
        for col in range(COL_PUESTOS_INICIO, COL_PUESTOS_FIN + 1):
            if col not in columnas_usadas:
                ws.cell(r, col).value = None


def parsear_multipart(body: bytes, content_type: str) -> dict:
    parser_input = b'Content-Type: ' + content_type.encode() + b'\r\n\r\n' + body
    msg = BytesParser(policy=default).parsebytes(parser_input)
    parts = {}
    for part in msg.iter_parts():
        nombre = part.get_param('name', header='content-disposition')
        if nombre:
            parts[nombre] = part.get_payload(decode=True)
    return parts


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            content_type = self.headers.get('Content-Type', '')
            body = self.rfile.read(content_length)

            parts = parsear_multipart(body, content_type)
            plantilla_bytes = parts.get('plantilla')
            data_bag_raw = parts.get('dataBag')

            if not plantilla_bytes or not data_bag_raw:
                self.send_response(400)
                self.end_headers()
                self.wfile.write(b'Faltan plantilla o dataBag')
                return

            data_bag = json.loads(data_bag_raw.decode('utf-8'))
            cabecera = data_bag.get('cabecera', {})
            items = data_bag.get('items', [])
            puestos = data_bag.get('puestos', [])

            if len(puestos) > MAX_PUESTOS:
                print(
                    f'[mpp.render-xlsx] {len(puestos)} puestos recibidos, '
                    f'capando a {MAX_PUESTOS} (límite de la plantilla)'
                )

            wb = load_workbook(BytesIO(plantilla_bytes))
            ws = wb[SHEET_NAME]
            reemplazar_placeholders(ws, cabecera)
            columnas_puestos = escribir_encabezados_puestos(ws, puestos)
            escribir_filas_mpp(ws, items, columnas_puestos)

            output = BytesIO()
            wb.save(output)
            output_bytes = output.getvalue()

            self.send_response(200)
            self.send_header(
                'Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            self.send_header('Content-Length', str(len(output_bytes)))
            self.end_headers()
            self.wfile.write(output_bytes)

        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({
                'error': str(e),
                'type': type(e).__name__
            }).encode('utf-8'))
